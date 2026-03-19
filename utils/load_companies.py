"""
Companies ETL loader for SuperTroopers.

Loads target company data from the Excel spreadsheet into the companies table.
Idempotent... uses UPSERT on company name so re-runs update existing rows.

Usage:
    python load_companies.py [--file PATH] [--dry-run]

Source:
    Templates/target_companies.xlsx (120 companies)
"""

import argparse
import os
import sys
from pathlib import Path

import openpyxl
import psycopg2
import psycopg2.extras

DEFAULT_EXCEL = str(
    Path(__file__).resolve().parent.parent.parent / "Templates" / "target_companies.xlsx"
)

# Column mapping: Excel header -> DB column
COLUMN_MAP = {
    "Company": "name",
    "Sector": "sector",
    "HQ/Location": "hq_location",
    "Size": "size",
    "Stage": "stage",
    "Fit Score": "fit_score",
    "Priority": "priority",
    "Target Role": "target_role",
    "Resume Variant": "resume_variant",
    "Key Differentiator": "key_differentiator",
    "Melbourne FL Relevant": "melbourne_relevant",
    "Comp Range": "comp_range",
    "Notes": "notes",
}

# DB columns in insert order (excluding id, created_at, updated_at)
DB_COLUMNS = [
    "name",
    "sector",
    "hq_location",
    "size",
    "stage",
    "fit_score",
    "priority",
    "target_role",
    "resume_variant",
    "key_differentiator",
    "melbourne_relevant",
    "comp_range",
    "notes",
]

UPSERT_SQL = """
INSERT INTO companies ({columns})
VALUES ({placeholders})
ON CONFLICT (name) DO UPDATE SET
    {updates},
    updated_at = NOW()
""".format(
    columns=", ".join(DB_COLUMNS),
    placeholders=", ".join(["%s"] * len(DB_COLUMNS)),
    updates=",\n    ".join(
        f"{col} = EXCLUDED.{col}" for col in DB_COLUMNS if col != "name"
    ),
)


from db_config import get_db_config


def get_connection():
    """Build a psycopg2 connection from env vars or .env file."""
    return psycopg2.connect(**get_db_config())


def ensure_unique_constraint(conn):
    """Make sure a unique constraint on companies.name exists for UPSERT."""
    with conn.cursor() as cur:
        cur.execute("""
            SELECT 1
            FROM pg_constraint
            WHERE conrelid = 'companies'::regclass
              AND contype = 'u'
              AND array_length(conkey, 1) = 1
              AND conkey[1] = (
                  SELECT attnum FROM pg_attribute
                  WHERE attrelid = 'companies'::regclass AND attname = 'name'
              )
        """)
        if cur.fetchone() is None:
            print("  Adding unique constraint on companies.name...")
            cur.execute(
                "ALTER TABLE companies ADD CONSTRAINT uq_companies_name UNIQUE (name)"
            )
            conn.commit()
            print("  Constraint added.")


def parse_excel(filepath):
    """Read the Excel file and return a list of row dicts keyed by DB column names."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    # Build header index
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_to_idx = {h: i for i, h in enumerate(headers) if h is not None}

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Skip completely empty rows
        if all(v is None for v in row):
            continue

        # Get company name... skip if missing
        company_idx = header_to_idx.get("Company")
        if company_idx is None or row[company_idx] is None:
            continue

        record = {}
        for excel_col, db_col in COLUMN_MAP.items():
            idx = header_to_idx.get(excel_col)
            if idx is not None and idx < len(row):
                val = row[idx]
                # Coerce fit_score to int
                if db_col == "fit_score" and val is not None:
                    try:
                        val = int(val)
                    except (ValueError, TypeError):
                        val = None
                # Coerce priority to single char
                if db_col == "priority" and val is not None:
                    val = str(val).strip()[:1].upper()
                # Trim strings
                if isinstance(val, str):
                    val = val.strip() or None
                record[db_col] = val
            else:
                record[db_col] = None

        rows.append(record)

    wb.close()
    return rows


def load_companies(rows, conn, dry_run=False):
    """Insert/update company rows into the database."""
    inserted = 0
    updated = 0
    skipped = 0

    with conn.cursor() as cur:
        for row in rows:
            values = [row.get(col) for col in DB_COLUMNS]
            name = row.get("name")

            if not name:
                skipped += 1
                continue

            if dry_run:
                print(f"  [DRY RUN] {name}")
                inserted += 1
                continue

            # Check if exists
            cur.execute("SELECT id FROM companies WHERE name = %s", (name,))
            exists = cur.fetchone() is not None

            cur.execute(UPSERT_SQL, values)

            if exists:
                updated += 1
                print(f"  Updated: {name}")
            else:
                inserted += 1
                print(f"  Inserted: {name}")

    if not dry_run:
        conn.commit()

    return inserted, updated, skipped


def build_parser(prog=None):
    parser = argparse.ArgumentParser(
        prog=prog,
        description="Load target companies from Excel into SuperTroopers DB",
    )
    parser.add_argument(
        "--file",
        default=DEFAULT_EXCEL,
        help=f"Path to target_companies.xlsx (default: {DEFAULT_EXCEL})",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and validate without writing to DB",
    )
    return parser


def main(argv=None):
    parser = build_parser()
    args = parser.parse_args(argv)

    excel_path = args.file
    if not os.path.exists(excel_path):
        print(f"ERROR: Excel file not found: {excel_path}", file=sys.stderr)
        sys.exit(1)

    print(f"Parsing {excel_path}...")
    rows = parse_excel(excel_path)
    print(f"  Found {len(rows)} companies")

    if not rows:
        print("Nothing to load.")
        return

    if args.dry_run:
        print("\nDry run... no DB changes will be made.")
        for row in rows:
            print(f"  {row.get('name')} | {row.get('sector')} | {row.get('priority')}")
        print(f"\n{len(rows)} companies would be loaded.")
        return

    print("\nConnecting to SuperTroopers DB...")
    conn = get_connection()
    print("  Connected.")

    print("Checking unique constraint on companies.name...")
    ensure_unique_constraint(conn)

    print(f"\nLoading {len(rows)} companies...")
    inserted, updated, skipped = load_companies(rows, conn)

    conn.close()

    print(f"\nDone. Inserted: {inserted}, Updated: {updated}, Skipped: {skipped}")


if __name__ == "__main__":
    main()
