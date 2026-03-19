"""
Applications ETL loader for SuperTroopers.

Loads application tracking data from two sources into the applications
and companies tables:
  1. Templates/application_tracker.xlsx (Excel tracker, primary source)
  2. Notes/APPLICATION_HISTORY.md (markdown with additional records)

Deduplicates by company+role, uses UPSERT for idempotency, and
auto-creates company entries where they don't exist.

Usage:
    python load_applications.py load [--dry-run]
    python load_applications.py load --excel-only
    python load_applications.py load --markdown-only
    python load_applications.py status
"""

import argparse
import os
import re
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

try:
    import psycopg2
    import psycopg2.extras
except ImportError:
    print(
        "ERROR: psycopg2-binary not installed. Run: pip install psycopg2-binary",
        file=sys.stderr,
    )
    sys.exit(1)


PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
EXCEL_PATH = PROJECT_ROOT / "Templates" / "application_tracker.xlsx"
MARKDOWN_PATH = PROJECT_ROOT / "Notes" / "APPLICATION_HISTORY.md"


from db_config import get_db_config


def get_db_connection():
    """Connect to the SuperTroopers PostgreSQL database."""
    return psycopg2.connect(**get_db_config())


def parse_date(value):
    """Parse a date from various formats. Returns a date string or None."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    if not s:
        return None
    # Try common formats
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def parse_excel(excel_path):
    """Parse the application tracker Excel file.

    Returns a list of dicts with keys matching the applications table columns.
    """
    if not excel_path.exists():
        print(f"  WARNING: Excel file not found at {excel_path}")
        return []

    wb = openpyxl.load_workbook(str(excel_path), read_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    # First row is the header
    header = [str(h).strip().lower().replace(" ", "_") if h else f"col_{i}"
              for i, h in enumerate(rows[0])]

    records = []
    for row in rows[1:]:
        if not row or all(v is None for v in row):
            continue
        data = dict(zip(header, row))
        record = {
            "company_name": (data.get("company") or "").strip() or None,
            "role": (data.get("role") or "").strip() or None,
            "date_applied": parse_date(data.get("date_applied")),
            "source": (data.get("source") or "").strip() or None,
            "status": (data.get("status") or "").strip() or None,
            "contact_name": (data.get("contact_name") or "").strip() or None,
            "contact_email": (data.get("contact_email") or "").strip() or None,
            "notes": (data.get("notes") or "").strip() or None,
            "jd_url": (data.get("jd_url") or "").strip() or None,
            "resume_version": (data.get("resume_version") or "").strip() or None,
            "origin": "excel",
        }
        # Skip rows without company or role
        if record["company_name"] and record["role"]:
            records.append(record)

    return records


def _extract_table_rows(lines, start_idx):
    """Extract data rows from a markdown table starting after the header.

    Given the line index of the ## heading, finds the header row, separator,
    and returns all data rows as lists of column strings.
    """
    rows = []
    i = start_idx + 1
    # Skip blank lines to find the header row
    while i < len(lines) and not lines[i].strip().startswith("|"):
        i += 1
    if i >= len(lines):
        return []
    # Skip header row
    i += 1
    # Skip separator row (|---|---|...)
    if i < len(lines) and re.match(r"\s*\|[-| :]+\|\s*$", lines[i]):
        i += 1
    # Collect data rows
    while i < len(lines):
        line = lines[i].strip()
        if not line.startswith("|"):
            break
        cols = [c.strip() for c in line.split("|")[1:-1]]
        if cols:
            rows.append(cols)
        i += 1
    return rows


def parse_markdown(md_path):
    """Parse APPLICATION_HISTORY.md for structured application records.

    Extracts from these markdown tables:
    - Rescinded Offers
    - Active Loops
    - Key Interviews Identified
    - Notable Rejections
    """
    if not md_path.exists():
        print(f"  WARNING: Markdown file not found at {md_path}")
        return []

    content = md_path.read_text(encoding="utf-8")
    lines = content.splitlines()
    records = []

    # Build index of section headings
    sections = {}
    for i, line in enumerate(lines):
        stripped = line.strip()
        if stripped.startswith("## "):
            sections[stripped] = i

    # --- Rescinded Offers ---
    for key, idx in sections.items():
        if "rescinded offers" in key.lower():
            for cols in _extract_table_rows(lines, idx):
                if len(cols) >= 3:
                    company = cols[0]
                    role = cols[1] if cols[1] != "TBD" else "Senior Engineering Leadership"
                    what_happened = cols[2]
                    records.append({
                        "company_name": company,
                        "role": role,
                        "date_applied": None,
                        "source": None,
                        "status": "Withdrawn",
                        "contact_name": None,
                        "contact_email": None,
                        "notes": f"Rescinded offer. {what_happened}" if what_happened else "Rescinded offer.",
                        "jd_url": None,
                        "resume_version": None,
                        "origin": "markdown",
                    })
            break

    # --- Active Loops ---
    for key, idx in sections.items():
        if "active loops" in key.lower():
            for cols in _extract_table_rows(lines, idx):
                if len(cols) >= 3:
                    company = cols[0]
                    status_raw = cols[1]
                    notes = cols[2]
                    records.append({
                        "company_name": company,
                        "role": "Senior Engineering Leadership",
                        "date_applied": None,
                        "source": None,
                        "status": "Interview",
                        "contact_name": None,
                        "contact_email": None,
                        "notes": f"{status_raw}. {notes}" if notes else status_raw,
                        "jd_url": None,
                        "resume_version": None,
                        "origin": "markdown",
                    })
            break

    # --- Key Interviews ---
    for key, idx in sections.items():
        if "key interviews" in key.lower():
            for cols in _extract_table_rows(lines, idx):
                if len(cols) >= 4:
                    date_str = cols[0]
                    company = cols[1]
                    role = cols[2]
                    outcome = cols[3]

                    date_applied = None
                    for fmt in ("%b %Y", "%B %Y"):
                        try:
                            date_applied = datetime.strptime(date_str, fmt).strftime("%Y-%m-01")
                            break
                        except ValueError:
                            continue

                    status = "Interview"
                    if "rejected" in outcome.lower():
                        status = "Rejected"

                    if role == "Unknown":
                        role = "Senior Engineering Leadership"

                    records.append({
                        "company_name": company,
                        "role": role,
                        "date_applied": date_applied,
                        "source": None,
                        "status": status,
                        "contact_name": None,
                        "contact_email": None,
                        "notes": f"Interview. Outcome: {outcome}",
                        "jd_url": None,
                        "resume_version": None,
                        "origin": "markdown",
                    })
            break

    # --- Notable Rejections ---
    for key, idx in sections.items():
        if "notable rejections" in key.lower():
            for cols in _extract_table_rows(lines, idx):
                if len(cols) >= 4:
                    date_str = cols[0]
                    company = cols[1]
                    role = cols[2]
                    reason = cols[3]

                    date_applied = None
                    for fmt in ("%b %Y", "%B %Y"):
                        try:
                            date_applied = datetime.strptime(date_str, fmt).strftime("%Y-%m-01")
                            break
                        except ValueError:
                            continue

                    if role in ("Unknown", "Manager"):
                        role = "Senior Engineering Leadership"

                    records.append({
                        "company_name": company,
                        "role": role,
                        "date_applied": date_applied,
                        "source": None,
                        "status": "Rejected",
                        "contact_name": None,
                        "contact_email": None,
                        "notes": reason,
                        "jd_url": None,
                        "resume_version": None,
                        "origin": "markdown",
                    })
            break

    return records


def deduplicate(records):
    """Deduplicate records by (company_name, role).

    Excel records take priority over markdown records. When both exist,
    we merge... markdown fields fill in blanks from the excel record.
    """
    seen = {}
    for rec in records:
        key = (
            (rec["company_name"] or "").lower().strip(),
            (rec["role"] or "").lower().strip(),
        )
        if key in seen:
            existing = seen[key]
            # Excel takes priority
            if existing["origin"] == "excel" and rec["origin"] == "markdown":
                # Merge: fill nulls in existing from markdown
                for field in rec:
                    if field == "origin":
                        continue
                    if existing.get(field) is None and rec.get(field) is not None:
                        existing[field] = rec[field]
            elif existing["origin"] == "markdown" and rec["origin"] == "excel":
                # Replace with excel, but keep any markdown-only data
                for field in existing:
                    if field == "origin":
                        continue
                    if rec.get(field) is None and existing.get(field) is not None:
                        rec[field] = existing[field]
                rec["origin"] = "excel"
                seen[key] = rec
            # If both same origin, keep the first one
        else:
            seen[key] = rec

    return list(seen.values())


def ensure_company(cur, company_name):
    """Ensure a company exists in the companies table. Returns the company id."""
    if not company_name:
        return None

    # Check if exists
    cur.execute(
        "SELECT id FROM companies WHERE LOWER(name) = LOWER(%s)",
        (company_name,),
    )
    row = cur.fetchone()
    if row:
        return row[0]

    # Insert new company
    cur.execute(
        "INSERT INTO companies (name) VALUES (%s) RETURNING id",
        (company_name,),
    )
    return cur.fetchone()[0]


def upsert_application(cur, record, company_id):
    """Upsert an application record.

    Uses company_name + role as the conflict key (no unique constraint in the
    schema, so we do a SELECT-then-INSERT/UPDATE pattern).
    """
    cur.execute(
        """SELECT id FROM applications
           WHERE LOWER(company_name) = LOWER(%s)
             AND LOWER(role) = LOWER(%s)""",
        (record["company_name"], record["role"]),
    )
    existing = cur.fetchone()

    if existing:
        # Update existing record... only overwrite non-null fields
        cur.execute(
            """UPDATE applications SET
                company_id = COALESCE(%s, company_id),
                date_applied = COALESCE(%s, date_applied),
                source = COALESCE(%s, source),
                status = COALESCE(%s, status),
                resume_version = COALESCE(%s, resume_version),
                jd_url = COALESCE(%s, jd_url),
                contact_name = COALESCE(%s, contact_name),
                contact_email = COALESCE(%s, contact_email),
                notes = COALESCE(%s, notes),
                last_status_change = NOW(),
                updated_at = NOW()
            WHERE id = %s""",
            (
                company_id,
                record["date_applied"],
                record["source"],
                record["status"],
                record["resume_version"],
                record["jd_url"],
                record["contact_name"],
                record["contact_email"],
                record["notes"],
                existing[0],
            ),
        )
        return "updated", existing[0]
    else:
        cur.execute(
            """INSERT INTO applications
                (company_id, company_name, role, date_applied, source, status,
                 resume_version, jd_url, contact_name, contact_email, notes,
                 last_status_change)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
            RETURNING id""",
            (
                company_id,
                record["company_name"],
                record["role"],
                record["date_applied"],
                record["source"],
                record["status"],
                record["resume_version"],
                record["jd_url"],
                record["contact_name"],
                record["contact_email"],
                record["notes"],
            ),
        )
        return "inserted", cur.fetchone()[0]


def load(dry_run=False, excel_only=False, markdown_only=False):
    """Main ETL load process."""
    all_records = []

    # --- Parse sources ---
    if not markdown_only:
        print(f"Parsing Excel: {EXCEL_PATH}")
        excel_records = parse_excel(EXCEL_PATH)
        print(f"  Found {len(excel_records)} records from Excel")
        all_records.extend(excel_records)

    if not excel_only:
        print(f"Parsing Markdown: {MARKDOWN_PATH}")
        md_records = parse_markdown(MARKDOWN_PATH)
        print(f"  Found {len(md_records)} records from Markdown")
        all_records.extend(md_records)

    if not all_records:
        print("No records found. Nothing to load.")
        return

    # --- Deduplicate ---
    print(f"\nDeduplicating {len(all_records)} total records...")
    records = deduplicate(all_records)
    print(f"  {len(records)} unique records after dedup")

    if dry_run:
        print("\n--- DRY RUN (no database changes) ---")
        for rec in records:
            print(
                f"  [{rec['origin']:>8}] {rec['company_name']:<30} "
                f"{rec['role']:<40} {rec['status']:<12} {rec['date_applied'] or 'no date'}"
            )
        print(f"\nDry run complete. {len(records)} records would be loaded.")
        return

    # --- Load into database ---
    print("\nConnecting to database...")
    conn = get_db_connection()
    cur = conn.cursor()

    inserted = 0
    updated = 0
    companies_created = 0
    errors = []

    try:
        for rec in records:
            try:
                # Ensure company exists
                company_id = ensure_company(cur, rec["company_name"])

                # Check if we just created a new company
                cur.execute("SELECT COUNT(*) FROM companies WHERE id = %s", (company_id,))
                # Track new companies by checking if it was just inserted
                # (we can't easily tell, so we'll count at the end)

                action, app_id = upsert_application(cur, rec, company_id)

                if action == "inserted":
                    inserted += 1
                    print(f"  + {rec['company_name']:<30} {rec['role']:<40} [{rec['status']}]")
                else:
                    updated += 1
                    print(f"  ~ {rec['company_name']:<30} {rec['role']:<40} [{rec['status']}]")

            except Exception as e:
                errors.append((rec["company_name"], rec["role"], str(e)))
                print(
                    f"  ! ERROR: {rec['company_name']} / {rec['role']}: {e}",
                    file=sys.stderr,
                )

        conn.commit()

        # Count companies we have
        cur.execute("SELECT COUNT(*) FROM companies")
        total_companies = cur.fetchone()[0]

        cur.execute("SELECT COUNT(*) FROM applications")
        total_apps = cur.fetchone()[0]

    finally:
        cur.close()
        conn.close()

    # --- Summary ---
    print(f"\n{'='*60}")
    print(f"Load complete.")
    print(f"  Inserted:  {inserted}")
    print(f"  Updated:   {updated}")
    print(f"  Errors:    {len(errors)}")
    print(f"  Total applications in DB: {total_apps}")
    print(f"  Total companies in DB:    {total_companies}")
    if errors:
        print(f"\nErrors:")
        for company, role, msg in errors:
            print(f"  - {company} / {role}: {msg}")


def status():
    """Show current application counts in the database."""
    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("SELECT COUNT(*) FROM applications")
        total_apps = cur.fetchone()[0]

        cur.execute("SELECT COUNT(*) FROM companies")
        total_companies = cur.fetchone()[0]

        cur.execute(
            "SELECT status, COUNT(*) FROM applications GROUP BY status ORDER BY COUNT(*) DESC"
        )
        statuses = cur.fetchall()

        cur.execute(
            "SELECT source, COUNT(*) FROM applications GROUP BY source ORDER BY COUNT(*) DESC"
        )
        sources = cur.fetchall()

        print(f"Applications: {total_apps}")
        print(f"Companies:    {total_companies}")
        print(f"\nBy Status:")
        for s, c in statuses:
            print(f"  {s or 'NULL':<20} {c}")
        print(f"\nBy Source:")
        for s, c in sources:
            print(f"  {s or 'NULL':<20} {c}")

    finally:
        cur.close()
        conn.close()


def build_parser(prog=None):
    parser = argparse.ArgumentParser(
        prog=prog,
        description="Load application tracking data into SuperTroopers database",
    )
    sub = parser.add_subparsers(dest="command", required=True)

    load_p = sub.add_parser("load", help="Parse sources and load into database")
    load_p.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and deduplicate but don't write to the database",
    )
    load_p.add_argument(
        "--excel-only",
        action="store_true",
        help="Only load from the Excel tracker",
    )
    load_p.add_argument(
        "--markdown-only",
        action="store_true",
        help="Only load from APPLICATION_HISTORY.md",
    )

    sub.add_parser("status", help="Show current application counts in the database")

    return parser


def main(argv=None):
    parser = build_parser()
    args = parser.parse_args(argv)

    if args.command == "load":
        load(
            dry_run=args.dry_run,
            excel_only=args.excel_only,
            markdown_only=args.markdown_only,
        )
    elif args.command == "status":
        status()


if __name__ == "__main__":
    main()
